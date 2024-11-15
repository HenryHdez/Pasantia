import math
import requests
import time
import json
import matplotlib.pyplot as plt
import random
from openpyxl import Workbook
from openpyxl import load_workbook

# Parámetros de movimiento
vx_script = 0.1
omega_script = 0.1
tolerancia_angular = 0.135
tolerancia_distancia = 0.05
intervalo_envio = 0.235
delta_t = intervalo_envio * 0.3159 + intervalo_envio

# Variables globales para la posición y orientación iniciales
x_actual = 0.0
y_actual = 0.0
theta_actual = 0.0
distancia_recorrida = 0.0
detecciones_obstaculos = 0
nombre_arch="metricas_robot_1.xlsx"
iteracion = 1  # Contador de iteraciones
potencia_robot = 88.8  # Potencia estimada en vatios (W), 12V 7.5Ah se consumen en 2 horas
tiempo_operativo = 0  # Tiempo en movimiento
ultimo_tiempo_ciclo = time.time()

def val_iniciales(x, y, t, vx, vt, ta, td):
    global x_actual, y_actual, theta_actual, vx_script, omega_script, tolerancia_angular, tolerancia_distancia
    vx_script = vx
    omega_script = vt
    tolerancia_angular = ta
    tolerancia_distancia = td
    x_actual = x
    y_actual = y
    theta_actual = t

def enviar_comandos_movimiento(ip, vx, vy, omega):
    url = f'http://{ip}/data/omnidrive'
    vel_vector = [vx, vy, omega]
    datos_json = json.dumps(vel_vector)
    try:
        headers = {'Content-Type': 'application/json'}
        response = requests.post(url, data=datos_json, headers=headers)
        if response.status_code != 200:
            print(f"Error: {response.text}")
    except requests.exceptions.RequestException as e:
        print(f'Error al enviar comandos de movimiento: {e}')

def leer_odometro(ip):
    url_odometry = f'http://{ip}/data/odometry'
    try:
        response = requests.get(url_odometry)
        if response.status_code == 200:
            odometro_datos = response.json()
            return odometro_datos[0], odometro_datos[1], odometro_datos[2], odometro_datos[3], odometro_datos[5]
        else:
            print(f"Error al leer el odómetro: {response.status_code}")
            return None
    except requests.exceptions.RequestException as e:
        print(f"Error al conectarse al odómetro: {e}")
        return None

def proximidad_robot(ip):
    url = f'http://{ip}/data/distancesensorarray'
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return response.json()
        else:
            print(f'Error al leer sensores: {response.status_code} - {response.text}')
            return [0] * 9
    except requests.exceptions.RequestException as e:
        print(f'Error al leer sensores: {e}')
        return [0] * 9

def integrar_velocidades(vx_odometro, omega_odometro, delta_t):
    global x_actual, y_actual, theta_actual, distancia_recorrida
    x_anterior, y_anterior = x_actual, y_actual
    x_actual += vx_odometro * delta_t * math.cos(theta_actual)
    y_actual += vx_odometro * delta_t * math.sin(theta_actual)
    distancia_recorrida += math.sqrt((x_actual - x_anterior) ** 2 + (y_actual - y_anterior) ** 2)
    theta_actual += omega_odometro * delta_t
    theta_actual = theta_actual % (2 * math.pi)

def calcular_angulo_objetivo(x_actual, y_actual, destino):
    delta_x = destino[0] - x_actual
    delta_y = destino[1] - y_actual
    return math.atan2(delta_y, delta_x)

def calcular_distancia(x_actual, y_actual, destino):
    return math.sqrt((destino[0] - x_actual) ** 2 + (destino[1] - y_actual) ** 2)

# Función principal de movimiento con gráficos en tiempo real y métricas adicionales
def mover_robot(ip, lista_destinos, tiempos_parada, Dimensiones="500x600+0+10", num_intermedios=1, nombre_arch="metricas_robot.xlsx"):
    global iteracion, detecciones_obstaculos, tiempo_operativo

    # Cargar o crear el archivo de Excel
    try:
        workbook = load_workbook(nombre_arch)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Métricas Robot"
        sheet.append(["Iteración", "Tiempo Total", "Distancia Recorrida", "Distancia al Objetivo", 
                      "Velocidad Lineal", "Orientación", "Detecciones de Obstáculos", 
                      "Disponibilidad (%)", "Tiempo de Ciclo (s)", "Consumo de Energía (Wh)", 
                      "Coordenada X", "Coordenada Y"])

    posiciones_robot = []
    tiempo_inicio = time.time()

    plt.switch_backend('TkAgg')
    fig, (ax_trayectoria, ax_sensores) = plt.subplots(2, 1, figsize=(5, 6))
    manager = plt.get_current_fig_manager()
    manager.window.wm_geometry(Dimensiones)

    # Configuración de la gráfica de trayectoria
    ax_trayectoria.set_xlim(-0.5, 2)
    ax_trayectoria.set_ylim(-0.5, 2)
    ax_trayectoria.grid(True)
    ax_trayectoria.set_title(f"Trayectoria del Robot en IP: {ip}")

    # Dibujar destinos y puntos intermedios
    destinos_x = [destino[0] for destino in lista_destinos]
    destinos_y = [destino[1] for destino in lista_destinos]
    ax_trayectoria.plot(destinos_x, destinos_y, 'ro', label="Destinos")

    for i in range(1, len(lista_destinos) - 1):
        if tiempos_parada[i] == 0:
            ax_trayectoria.plot(lista_destinos[i][0], lista_destinos[i][1], 'o', color='violet', label="Punto Intermedio")

    # Elementos gráficos para la trayectoria y orientación
    trayectoria_robot, = ax_trayectoria.plot([], [], 'b-', lw=2, label="Trayectoria Robot")
    robot_actual, = ax_trayectoria.plot([], [], 'go', label="Posición Robot")
    orientacion_robot, = ax_trayectoria.plot([], [], 'r-', lw=2, label="Orientación")

    # Gráfico de sensores de proximidad
    ax_sensores.set_ylim(0, 1)
    ax_sensores.set_title("Sensores de Proximidad")
    ax_sensores.set_xlabel("Sensor")
    ax_sensores.set_ylabel("Distancia")
    barras_sensores = ax_sensores.bar(range(9), [0]*9, color='c')

    for i, destino in enumerate(lista_destinos):
        print(f"Moviendo hacia el destino: {destino}")
        tiempo_inicio_ciclo = time.time()

        while True:
            posicion = leer_odometro(ip)
            sensores = proximidad_robot(ip)
            if posicion:
                x_odometro, y_odometro, theta_odometro, vx_odometro, omega_odometro = posicion
                angulo_objetivo = calcular_angulo_objetivo(x_actual, y_actual, destino)
                delta_angulo = (angulo_objetivo - theta_actual + math.pi) % (2 * math.pi) - math.pi
                distancia = calcular_distancia(x_actual, y_actual, destino)
                if sensores[0] < 0.12:
                    detecciones_obstaculos += 1
                    print("Obstáculo detectado! Realizando maniobra de esquiva.")
                    continue
                elif abs(delta_angulo) > tolerancia_angular:
                    omega = omega_script if delta_angulo > 0 else -omega_script
                    enviar_comandos_movimiento(ip, 0, 0, omega)
                elif distancia > tolerancia_distancia:
                    enviar_comandos_movimiento(ip, vx_script, 0, 0)
                    tiempo_operativo += delta_t
                else:
                    print(f"Llegó al destino: {destino}")
                    enviar_comandos_movimiento(ip, 0, 0, 0)
                    
                    if i < len(tiempos_parada) and tiempos_parada[i] > 0:
                        time.sleep(tiempos_parada[i])
                    break

                integrar_velocidades(vx_odometro, omega_odometro, delta_t)
                posiciones_robot.append([x_actual, y_actual])

                # Calcular métricas adicionales
                tiempo_total = time.time() - tiempo_inicio
                distancia_al_objetivo = calcular_distancia(x_actual, y_actual, destino)
                velocidad_lineal_real = vx_odometro
                orientacion = theta_actual
                disponibilidad = (tiempo_operativo / tiempo_total) * 100
                tiempo_ciclo = time.time() - tiempo_inicio_ciclo
                consumo_energia = potencia_robot * tiempo_operativo / 3600

                # Graficar trayectoria
                trayectorias_x = [pos[0] for pos in posiciones_robot]
                trayectorias_y = [pos[1] for pos in posiciones_robot]
                trayectoria_robot.set_data(trayectorias_x, trayectorias_y)
                
                # Graficar posición y orientación
                robot_actual.set_data([x_actual], [y_actual])
                longitud_flecha = 0.2
                orientacion_x = [x_actual, x_actual + longitud_flecha * math.cos(theta_actual)]
                orientacion_y = [y_actual, y_actual + longitud_flecha * math.sin(theta_actual)]
                orientacion_robot.set_data(orientacion_x, orientacion_y)

                # Actualizar los gráficos de sensores
                for j, barra in enumerate(barras_sensores):
                    barra.set_height(sensores[j])

                plt.draw()
                plt.pause(0.05)

                # Guardar métricas en Excel
                sheet.append([iteracion, tiempo_total, distancia_recorrida, distancia_al_objetivo, 
                              velocidad_lineal_real, orientacion, detecciones_obstaculos,
                              disponibilidad, tiempo_ciclo, consumo_energia,
                              x_actual, y_actual])
                
                # Guardar el archivo de Excel
                workbook.save(nombre_arch)
                iteracion += 1

if __name__ == "__main__":
    ip_robot = "192.168.1.101"
    lista_destinos = [(0.0, 0.0), (0.25, 0.75), (0.0, 1.5), (0.5, 1.5), (1.0, 1.5), (0.75, 0.75), (1.0, 0.0), (0.0, 0.0)]
    tiempos_parada = [3, 5, 3, 6, 8, 5, 6, 8]
    for i in range(3):
        mover_robot(ip_robot, lista_destinos, tiempos_parada)
